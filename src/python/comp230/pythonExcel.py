from openpyxl import load_workbook
from datetime import datetime

def processSheet(sheetInput): # sheetInput is <Worksheet "customer"> <Worksheet "product"> etc...
    loadSheetToDict = dict()
    # sheet[1] is equal to a header row of a sheet if any
    header_row = sheetInput[1]
    # print(f"sheetInput {sheetInput}")
    for row in sheetInput:
        if row != header_row:
            rowDict = dict()
            index = 0
            for cell in row:
                # header_row[index].value: index=0 is A1 and value is customer_id, etc..
                rowDict[header_row[index].value] = cell.value
                index += 1
            # row[0]: use the first column as a dictionary key
            loadSheetToDict[row[0].value] = rowDict
    return loadSheetToDict

def result(file):
    workbook = load_workbook(file)
    worksheets = workbook.sheetnames
    index = 0
    # worksheet: return list of string '...' worksheet names (string) such as customer, product, etc..
    for worksheet in worksheets:
        # workbook[worksheet] - workbook['customer'] returns <Worksheet "customer">
        worksheets[index] = processSheet(workbook[worksheet])
        index += 1
    return worksheets

def getProductName(data, prodid):
    listOfKeys = list()
    prodResult = ""
    storedData = data
    listOfItems = storedData[1]  # Product dictionary
    for item in listOfItems.values():
        if item['prod_id'] == prodid:
            # prodResult = f"Product Name: {item['prod_name']}, Unit Price: {item['prod_price']}"
            prodResult = f"{item['prod_name']}\t\t{item['prod_price']}"
            return prodResult

def getCustomerName(data, custid):
    listOfKeys = list()
    customerResult = ""
    storedData = data
    listOfItems = storedData[0]  # Customer dictionary
    for item in listOfItems.values():
        if item['customer_id'] == custid:
            customerResult = f"Customer Name: {item['first_name']} {item['last_name']}"
            return customerResult

def report(invoiceInput):
    invInput=""
    message=""
    if invoiceInput==None:
        invInput=input("please enter an invoice #: ")
    else:
        invInput=invoiceInput
    # listOfKeys = list()
    storedData = result("./excelFile/sampledata.xlsx")
    getInvID = dict()
    getInvoiceInfo = []
    custName = ""
    cust_id = 0
    invoiceInfo = ""
    amount=0
    info="" # to get the result returned from Invoice dictionary
    valid = False # to validate an invoice No.
    # to validate invoce # from user input
    for checkID in storedData[2].values(): # storedData[2]: data from invoice dictionary
        if checkID['invoice_id'] == int(invInput):
            valid = True
    if valid == False:
        message= "Invoice No. does not exist."
    else:
        for inv_id in storedData[2].values():
            if inv_id['invoice_id'] == int(invInput):
                cust_id = inv_id['customer_id']
                custName = getCustomerName(storedData, cust_id)
                info=f"Invoice No. {inv_id['invoice_id']}. Invoice Date: {inv_id['invoice_date']}\n{custName}\n"
                str1="Sales_Amt:".ljust(22,' ')
                str2="GST:".ljust(23,' ')
                str3="Total Sales:".ljust(22,' ')
                amount=f"{str1} {format(inv_id['sales_amount'],'.2f')}\n{str2} {format(inv_id['gst'],'.2f')}\n{str3} {float(format(inv_id['total_amount'],'.2f'))}\n"
        for item in storedData[3].values(): # storedData[3] data from Invoice_Detail dictionary
            if item['invoice_id'] == int(invInput):
                getInvoiceInfo.append(item)
        for prodID in getInvoiceInfo:  
            invoiceInfo += f"{getProductName(storedData,prodID['prod_id']).rjust(25,' ')}\t\t{str(prodID['quantity']).rjust(2,' ')}\n"
        title=f"PRODUCT NAME:\t\tUNIT PRICE:\t\tQTY:"
        f = open("/home/danny/code/cohort3/src/python/comp230/invoice.txt", "w")
        f.write(f"{info}\n{title}\n{invoiceInfo}\n{amount}") # write Invoice to txt
        # return f"{info}\n{title}\n{invoiceInfo}\n{amount}"
        message= "Invoice generated"
    return message

# Validating sheets:

def getColNames(sheetInput):
    colNames=[]
    for col in sheetInput.iter_cols(1, sheetInput.max_column):
        colNames.append(col[0].value) # col[0].value: title value of each columns such as customer_id, first_name, etc..
    return colNames
def validateFile(fileInput): # STEP 1-to validate correct FILE extension, COLUMNS name, order of SHEETS
    if fileInput.endswith('.xlsx') or fileInput.endswith('.XLSX'): # only allow .xlsx OR XLSX file at this time
        workbook = load_workbook(fileInput)
        worksheet = workbook.sheetnames
        numOfSheets=len(worksheet)
        catchError=""
        # sheetsInorder=False # to validate if all sheets in specific order
        if  worksheet==['customer', 'product', 'invoice', 'invoice_detail'] and numOfSheets==4:
            correct=0
            holdListOfColNames=[]
            # f = open("./errorLog/errLog-validateBySheets.txt", "w")
            for s in worksheet: 
                if s=='customer':
                    colFromCustomer=getColNames(workbook[s])
                    holdListOfColNames.append(colFromCustomer)
                    expectCol=['customer_id', 'first_name', 'last_name', 'address', 'phone', 'email']
                    if holdListOfColNames[0]==expectCol:
                        correct+=1
                    else:
                        catchError+=f"- First sheet expected 'customer' and 6 columns in order {expectCol}\n"
                elif s=='product':
                    colFromCustomer=getColNames(workbook[s])
                    holdListOfColNames.append(colFromCustomer)
                    expectCol=['prod_id', 'prod_name', 'prod_price']
                    if holdListOfColNames[1]==expectCol:
                        correct+=1
                    else:
                        catchError+=f"- Second sheet expected 'product' and 3 columns in order {expectCol}\n"
                elif s=='invoice':
                    colFromCustomer=getColNames(workbook[s])
                    holdListOfColNames.append(colFromCustomer)
                    expectCol=['invoice_id', 'customer_id', 'invoice_date', 'sales_amount', 'gst', 'total_amount']
                    if holdListOfColNames[2]==expectCol:
                        correct+=1
                    else:
                        catchError+=f"- Third sheet expected 'invoice' and 6 columns in order {expectCol}\n"
                else:
                    colFromCustomer=getColNames(workbook[s])
                    holdListOfColNames.append(colFromCustomer)
                    expectCol=['detail_id', 'invoice_id', 'prod_id', 'quantity']
                    if holdListOfColNames[3]==expectCol:
                        correct+=1
                    else:
                        catchError+=f"- Last sheet expected 'invoice_detail' and 4 columns in order {expectCol}\n"
            if correct!=4:
                print('Please refer to errorLog_validateBySheets.txt')
                f = open("./errorLog/errLog-validateBySheets.txt", "w")
                f.write(f"Error @ {datetime.now()}\n{catchError}")
                f.close()
            return correct
        else:
            return "File is incorrect format. The file must contain 4 sheets and be in correct order: customer, product, invoice, invoice_detail"
    else:
        return "Incorrect File extension. It Only accepts .xlsx"

def validateContent(fileName): # STEP 2 - validate individual cells data - use this after running VALIDATEFILE()
    fileChecked=validateFile(fileName)
    if fileChecked==4:
        workbook=load_workbook(fileName)
        # f = open("errLog-CellDataValidation.txt", "w")
        anyErr=False
        catchErrLine=""
        for strname in workbook.sheetnames:
            if strname=='customer':
                err=""
                for s in workbook[strname]:
                    header=workbook[strname][1][0].value
                    if s[0].value != header: # check values in 1st column which is customer_id  
                        if type(s[0].value)!=int:
                            err+="True"
                            catchErrLine+=f"\tError at {s[0]} - Value is '{s[0].value}'. It must be > 0 and can't be empty.\n"
                if err.find("True") ==-1: # -1 is no error found
                    print (f"Data from ({strname}) sheet has been checked and Ready to merge.")
                else:
                    anyErr=True
                    print("Please refer to errLog-CellDataValidation.txt for more error information.")
                    # f.write(f"Error @ {datetime.now()}\n{catchErrLine}")
            elif strname=='product':
                err=""
                for s in workbook[strname]:
                    prodID=workbook[strname][1][0].value
                    if s[0].value != prodID: # check values in 1st column which is product_id
                        if type(s[0].value)!=int:
                            err+="True"
                            catchErrLine+=f"\tError at {s[0]} - Value is '{s[0].value}'. It must be > 0 and can't be empty.\n"
                    prodPrice=workbook[strname][1][2].value
                    if s[2].value != prodPrice: # check values in 3rd column which is prod_price
                        if type(s[2].value)==str:
                            err+="True"
                            catchErrLine+=f"\tError at {s[2]} - Value is '{s[2].value}'. It must be > 0 and can't be empty.\n"
                if err.find("True")==-1:
                    print (f"Data from ({strname}) sheet has been checked and Ready to merge.")
                else:
                    anyErr=True
                    print("Please refer to errLog-CellDataValidation.txt for more error information.")
                    # f.write(f"Error @ {datetime.now()}\n{catchErrLine}")
            elif strname=='invoice':
                err=""
                for s in workbook[strname]:
                    InvoiceID=workbook[strname][1][0].value
                    if s[0].value != InvoiceID: # check values in 1st column which is invoice_id
                        if type(s[0].value)!=int:
                            err+="True"
                            catchErrLine+=f"\tError at {s[0]} - Value is '{s[0].value}'. It must be > 0 and can't be empty.\n"
                    cusID=workbook[strname][1][1].value
                    if s[1].value != cusID: # check values in 2st column which is customer_id
                        if type(s[1].value)!=int:
                            err+="True"
                            catchErrLine+=f"\tError at {s[1]} - Value is '{s[1].value}'. It must be > 0 and can't be empty.\n"

                    invoiceDate=workbook[strname][1][2].value
                    if s[2].value != invoiceDate: # check values in 3st column which is invoice_date
                        checkDate=str(s[2].value)
                        try:
                            datetime.strptime(checkDate, '%Y-%m-%d 00:00:00')
                        except ValueError:
                            err+="True"
                            a=f"\tError at {s[2]} - Value is '{s[2].value}'. It must be YYYY-MM-DD and can't be empty.\n"
                            catchErrLine+=a

                    combineDEF=[s[3].value,s[4].value,s[5].value]
                    if combineDEF !=['sales_amount', 'GST(5%)', 'total_amount']:
                        if type(s[3].value)==str:
                            err+='True'
                            catchErrLine+=f"\tError at {s[3]} - Value is ({s[3].value}). It must be a number and can't be empty.\n"
                        if type(s[4].value)==str:
                            err+="True"
                            catchErrLine+=f"\tError at {s[4]} - Value is ({s[4].value}). It must be a number and can't be empty.\n"
                        if type(s[5].value)==str:
                            err+="True"
                            catchErrLine+=f"\tError at {s[5]} - Value is ({s[5].value}). It must be a number and can't be empty.\n"
                if err.find("True")==-1:
                    print (f"Data from ({strname}) sheet has been checked and Ready to merge.")
                else:
                    anyErr=True
                    print("Please refer to errLog-CellDataValidation.txt for more error information.")
                    # f.write(f"Error @ {datetime.now()}\n{catchErrLine}")
            else: # strname=='invoice_detail':
                err=""
                for s in workbook[strname]:
                    combineABCD=[s[0].value,s[1].value,s[2].value,s[3].value]
                    if combineABCD !=['detail_id', 'invoice_id', 'prod_id', 'quantity']:
                        if type(s[0].value)!=int:
                            err+="True"
                            catchErrLine+=f"\tError at {s[0]} - Value is ({s[0].value}). It must be > 0 and can't be empty.\n"
                        if type(s[1].value)!=int:
                            err+="True"
                            catchErrLine+=f"\tError at {s[1]} - Value is ({s[1].value}). It must be > 0 and can't be empty.\n"
                        if type(s[2].value)!=int:
                            err+="True"
                            catchErrLine+=f"\tError at {s[2]} - Value is ({s[2].value}). It must be > 0 and can't be empty.\n"
                        if type(s[3].value)!=int:
                            err+="True"
                            catchErrLine+=f"\tError at {s[3]} - Value is ({s[3].value}). It must be > 0 and can't be empty.\n"
                if err.find("True")==-1:
                    print (f"Data from ({strname}) sheet has been checked and Ready to merge.")
                else:
                    anyErr=True
                    print("Please refer to errLog-CellDataValidation.txt for more error information.")
        if anyErr==True:
            f = open("./errorLog/errLog-CellDataValidation.txt", "w")
            f.write(f"Error @ {datetime.now()}\n{catchErrLine}")
            f.close()
    else:
        print("The file does not match the destination file format. Please correct before merging.")
 
# validateContent('./excelFile/test.xlsx')

def mergeData(template,set1):
    templateWorkbook=load_workbook(template)
    wb1=load_workbook(set1)
    # tempCustomer=templateWorkbook.get_sheet_by_name('customer')
    try:
        for sName in templateWorkbook.sheetnames:
            print("Data has been processing...")
            tempCustomer=templateWorkbook.get_sheet_by_name(sName)
            srcCustomer=wb1.get_sheet_by_name(sName)
            for r in range(2,srcCustomer.max_row+1):
                m=tempCustomer.max_row
                for c in range(1,srcCustomer.max_column+1):
                    tempCustomer.cell(row=m+1, column=c).value = srcCustomer.cell(row=r, column=c).value
        templateWorkbook.save('./excelFile/template.xlsx')
        print("Data has been added to the destination.")
    except ValueError:
        raise ValueError


    # for sName in templateWorkbook.sheetnames:
    #     if sName=='customer':
    #         tempCustomer=templateWorkbook.get_sheet_by_name(sName)
    #         srcCustomer=wb1.get_sheet_by_name(sName)
    #         for r in range(2,srcCustomer.max_row+1):
    #             m=tempCustomer.max_row
    #             for c in range(1,srcCustomer.max_column+1):
    #                 tempCustomer.cell(row=m+1, column=c).value = srcCustomer.cell(row=r, column=c).value
    # templateWorkbook.save('./excelFile/template.xlsx')


