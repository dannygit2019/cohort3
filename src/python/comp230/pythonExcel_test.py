import pytest
import pythonExcel
from openpyxl import load_workbook


def test_workSheetInMemory():
    workbook = load_workbook(filename="./excelFile/sampledata.xlsx")
    # load CUSTOMER data
    sheetIn = workbook['customer']
    assert pythonExcel.processSheet(sheetIn)=={1: {'customer_id': 1, 'first_name': 'Rachel', 'last_name': 'Smith', 'address': '123 Riverbend SE', 'phone': '403-324-3455', 'email': 'rachel@telus.ca'}, 2: {'customer_id': 2, 'first_name': 'Bobby', 'last_name': 'Smith', 'address': '435 Crest Place NE', 'phone': '587-321-4564', 'email': 'bobby@gmail.com'}, 3: {'customer_id': 3, 'first_name': 'Rebecca', 'last_name': 'Wang', 'address': '231 Crest Place NE', 'phone': '403-132-5798', 'email': 'rebecca@hotmail.com'}, 4: {'customer_id': 4, 'first_name': 'Tim', 'last_name': 'Cormier', 'address': '435 Redstone NE', 'phone': '403-432-4532', 'email': 'tim@gmail.com'}, 5: {'customer_id': 5, 'first_name': 'Phil', 'last_name': 'Murray', 'address': '568 NoseHill', 'phone': '587-091-0934', 'email': 'phil@shaw.ca'}, 6: {'customer_id': 6, 'first_name': 'Shane', 'last_name': 'James', 'address': '987 Brightton SE', 'phone': '587-765-7898', 'email': 'shane@gmail.com'}, 7: {'customer_id': 7, 'first_name': 'John', 'last_name': 'Sandhu', 'address': '356 Willow Park', 'phone': '587-321-2343', 'email': 'john@telus.ca'}, 8: {'customer_id': 8, 'first_name': 'Dave', 'last_name': 'Hynes', 'address': '1213 Brightton SE', 'phone': '403-554-2321', 'email': 'dave@hotmai.com'}, 9: {'customer_id': 9, 'first_name': 'Cristina', 'last_name': 'Davey', 'address': '451 Kensington', 'phone': '587-090-2931', 'email': 'cristina@gmail.com'}, 10: {'customer_id': 10, 'first_name': 'Dave', 'last_name': 'James', 'address': '189 SunnySide', 'phone': '403-432-7094', 'email': 'dave@shaw.ca'}}
    # load PRODUCT data
    sheetIn = workbook['product']
    assert pythonExcel.processSheet(sheetIn)=={1: {'prod_id': 1, 'prod_name': 'Dishwasher', 'prod_price': 599.99}, 2: {'prod_id': 2, 'prod_name': 'Wine Cooler', 'prod_price': 199.99}, 3: {'prod_id': 3, 'prod_name': 'Microwave', 'prod_price': 159.99}, 4: {'prod_id': 4, 'prod_name': 'Mini Oven', 'prod_price': 299.99}, 5: {'prod_id': 5, 'prod_name': 'Electric Stove', 'prod_price': 899.99}, 6: {'prod_id': 6, 'prod_name': 'Gas Stove', 'prod_price': 1119.99}, 7: {'prod_id': 7, 'prod_name': 'Blender', 'prod_price': 129.99}, 8: {'prod_id': 8, 'prod_name': 'Portable Heater', 'prod_price': 119.99}, 9: {'prod_id': 9, 'prod_name': 'Dining aTable Set', 'prod_price': 899.99}, 10: {'prod_id': 10, 'prod_name': 'Washer', 'prod_price': 999.99}}
    # load INVOICE data
    sheetIn = workbook['invoice']
    assert pythonExcel.processSheet(sheetIn)=={1: {'invoice_id': 1, 'customer_id': 1, 'invoice_date': '12-01-2019', 'sales_amount': 2499.97, 'gst': 124.9985, 'total_amount': 2624.9685}, 2: {'invoice_id': 2, 'customer_id': 2, 'invoice_date': '12-05-2019', 'sales_amount': 2019.97, 'gst': 100.9985, 'total_amount': 4745.937}, 3: {'invoice_id': 3, 'customer_id': 3, 'invoice_date': '12-09-2019', 'sales_amount': 1419.97, 'gst': 70.9985, 'total_amount': 3611.937}}
    # load INVOICE_DETAIL data
    sheetIn = workbook['invoice_detail']
    assert pythonExcel.processSheet(sheetIn)=={1: {'detail_id': 1, 'invoice_id': 1, 'prod_id': 1, 'quantity': 1}, 2: {'detail_id': 2, 'invoice_id': 1, 'prod_id': 10, 'quantity': 1}, 3: {'detail_id': 3, 'invoice_id': 1, 'prod_id': 9, 'quantity': 1}, 4: {'detail_id': 4, 'invoice_id': 2, 'prod_id': 8, 'quantity': 1}, 5: {'detail_id': 5, 'invoice_id': 2, 'prod_id': 9, 'quantity': 1}, 6: {'detail_id': 6, 'invoice_id': 2, 'prod_id': 10, 'quantity': 1}, 7: {'detail_id': 7, 'invoice_id': 3, 'prod_id': 4, 'quantity': 1}, 8: {'detail_id': 8, 'invoice_id': 3, 'prod_id': 8, 'quantity': 1}, 9: {'detail_id': 9, 'invoice_id': 3, 'prod_id': 10, 'quantity': 1}}

def testresult():
    # to display customer info
    assert pythonExcel.result("./excelFile/sampledata.xlsx")[0]=={1: {'customer_id': 1, 'first_name': 'Rachel', 'last_name': 'Smith', 'address': '123 Riverbend SE', 'phone': '403-324-3455', 'email': 'rachel@telus.ca'}, 2: {'customer_id': 2, 'first_name': 'Bobby', 'last_name': 'Smith', 'address': '435 Crest Place NE', 'phone': '587-321-4564', 'email': 'bobby@gmail.com'}, 3: {'customer_id': 3, 'first_name': 'Rebecca', 'last_name': 'Wang', 'address': '231 Crest Place NE', 'phone': '403-132-5798', 'email': 'rebecca@hotmail.com'}, 4: {'customer_id': 4, 'first_name': 'Tim', 'last_name': 'Cormier', 'address': '435 Redstone NE', 'phone': '403-432-4532', 'email': 'tim@gmail.com'}, 5: {'customer_id': 5, 'first_name': 'Phil', 'last_name': 'Murray', 'address': '568 NoseHill', 'phone': '587-091-0934', 'email': 'phil@shaw.ca'}, 6: {'customer_id': 6, 'first_name': 'Shane', 'last_name': 'James', 'address': '987 Brightton SE', 'phone': '587-765-7898', 'email': 'shane@gmail.com'}, 7: {'customer_id': 7, 'first_name': 'John', 'last_name': 'Sandhu', 'address': '356 Willow Park', 'phone': '587-321-2343', 'email': 'john@telus.ca'}, 8: {'customer_id': 8, 'first_name': 'Dave', 'last_name': 'Hynes', 'address': '1213 Brightton SE', 'phone': '403-554-2321', 'email': 'dave@hotmai.com'}, 9: {'customer_id': 9, 'first_name': 'Cristina', 'last_name': 'Davey', 'address': '451 Kensington', 'phone': '587-090-2931', 'email': 'cristina@gmail.com'}, 10: {'customer_id': 10, 'first_name': 'Dave', 'last_name': 'James', 'address': '189 SunnySide', 'phone': '403-432-7094', 'email': 'dave@shaw.ca'}}
    # to display product info
    assert pythonExcel.result("./excelFile/sampledata.xlsx")[1]=={1: {'prod_id': 1, 'prod_name': 'Dishwasher', 'prod_price': 599.99}, 2: {'prod_id': 2, 'prod_name': 'Wine Cooler', 'prod_price': 199.99}, 3: {'prod_id': 3, 'prod_name': 'Microwave', 'prod_price': 159.99}, 4: {'prod_id': 4, 'prod_name': 'Mini Oven', 'prod_price': 299.99}, 5: {'prod_id': 5, 'prod_name': 'Electric Stove', 'prod_price': 899.99}, 6: {'prod_id': 6, 'prod_name': 'Gas Stove', 'prod_price': 1119.99}, 7: {'prod_id': 7, 'prod_name': 'Blender', 'prod_price': 129.99}, 8: {'prod_id': 8, 'prod_name': 'Portable Heater', 'prod_price': 119.99}, 9: {'prod_id': 9, 'prod_name': 'Dining aTable Set', 'prod_price': 899.99}, 10: {'prod_id': 10, 'prod_name': 'Washer', 'prod_price': 999.99}}
    # to display invoice info
    assert pythonExcel.result("./excelFile/sampledata.xlsx")[2]=={1: {'invoice_id': 1, 'customer_id': 1, 'invoice_date': '12-01-2019', 'sales_amount': 2499.97, 'gst': 124.9985, 'total_amount': 2624.9685}, 2: {'invoice_id': 2, 'customer_id': 2, 'invoice_date': '12-05-2019', 'sales_amount': 2019.97, 'gst': 100.9985, 'total_amount': 4745.937}, 3: {'invoice_id': 3, 'customer_id': 3, 'invoice_date': '12-09-2019', 'sales_amount': 1419.97, 'gst': 70.9985, 'total_amount': 3611.937}}
    # to display invoice_detail info
    assert pythonExcel.result("./excelFile/sampledata.xlsx")[3]=={1: {'detail_id': 1, 'invoice_id': 1, 'prod_id': 1, 'quantity': 1}, 2: {'detail_id': 2, 'invoice_id': 1, 'prod_id': 10, 'quantity': 1}, 3: {'detail_id': 3, 'invoice_id': 1, 'prod_id': 9, 'quantity': 1}, 4: {'detail_id': 4, 'invoice_id': 2, 'prod_id': 8, 'quantity': 1}, 5: {'detail_id': 5, 'invoice_id': 2, 'prod_id': 9, 'quantity': 1}, 6: {'detail_id': 6, 'invoice_id': 2, 'prod_id': 10, 'quantity': 1}, 7: {'detail_id': 7, 'invoice_id': 3, 'prod_id': 4, 'quantity': 1}, 8: {'detail_id': 8, 'invoice_id': 3, 'prod_id': 8, 'quantity': 1}, 9: {'detail_id': 9, 'invoice_id': 3, 'prod_id': 10, 'quantity': 1}}

def test_getProductname():
    dataInput=pythonExcel.result("./excelFile/sampledata.xlsx")
    assert pythonExcel.getProductName(dataInput,1)=="Dishwasher\t\t599.99"
def test_getCustomername():
    dataInput=pythonExcel.result("./excelFile/sampledata.xlsx")
    assert pythonExcel.getCustomerName(dataInput,1)=="Customer Name: Rachel Smith"
def testReport():
    assert pythonExcel.report(51)=="Invoice No. does not exist."
    assert pythonExcel.report(3)=="Invoice generated"

# Validation exercise
def testgetColNames(): # getColNames(sheetInput)
    workbook = load_workbook(filename="./excelFile/sampledata.xlsx")
    sheetIn = workbook['customer']
    assert pythonExcel.getColNames(sheetIn)==['customer_id', 'first_name', 'last_name', 'address', 'phone', 'email']
    sheetIn = workbook['product']
    assert pythonExcel.getColNames(sheetIn)==['prod_id', 'prod_name', 'prod_price']

def testvalidateFile():
    assert pythonExcel.validateFile('./excelFile/sampledata.csv')=="Incorrect File extension. It Only accepts .xlsx"
    assert pythonExcel.validateFile('./excelFile/makeTestFail.xlsx')=="File is incorrect format. The file must contain 4 sheets and be in correct order: customer, product, invoice, invoice_detail"
    assert pythonExcel.validateFile('./excelFile/individualSheetFail.xlsx')==0 # each sheets have an extra column
    assert pythonExcel.validateFile('./excelFile/test.xlsx')==4 # all sheets passed

def validateContent():
    pythonExcel.validateContent('./excelFile/test.xlsx')

# def testMerging():
    # pythonExcel.mergeData('./excelFile/template.xlsx','./excelFile/dataset1.xlsx')
    # pythonExcel.mergeData('./excelFile/template.xlsx','./excelFile/dataset2.xlsx')
    # pythonExcel.mergeData('./excelFile/template.xlsx','./excelFile/dataset3.xlsx')