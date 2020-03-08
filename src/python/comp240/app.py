from flask import Flask, jsonify, request, render_template
from openpyxl import load_workbook
import json
from flask_cors import CORS


app=Flask(__name__)
CORS(app, supports_credentials=True)
# @app.route('/')
# def home():
#     return "Hello, world"

data={}

def processSheet(sheetInput): # sheetInput is <Worksheet "customer"> <Worksheet "product"> etc...
    # loadSheetToDict = []
    loadSheetToDict = dict()
    header_row = sheetInput[1]
    for row in sheetInput:
        if row != header_row:
            rowDict = dict()
            index = 0
            for cell in row:
                rowDict[header_row[index].value] = cell.value
                index += 1
            # print(rowDict)   
            loadSheetToDict[row[0].value] = rowDict
            # loadSheetToDict.append(rowDict)
    # print(loadSheetToDict) 
    return loadSheetToDict

def result(file):
    workbook = load_workbook(file)
    worksheets = workbook.sheetnames
    index = 0
    allSheets=[]
    for worksheet in worksheets:
        allSheets.append(processSheet(workbook[worksheet]))
    return allSheets

getResult=result('dataset.xlsx')

# @app.route('/all',methods=['GET'])
# def displayAll():
#     return jsonify(getResult)

# @app.route('/customers', methods=['GET'])
# def displayCustomer():
#     return jsonify(getResult[0])

# @app.route('/products')
# def displayProduct():
#     return jsonify(getResult[1])

# @app.route('/invoices')
# def displayInvoices():
#     return jsonify(getResult[2])

# @app.route('/invoice_detail')
# def displayInvDetail():
#     return jsonify(getResult[3])

@app.route('/display') # flask templates
def displayToHTML():
    cus=getResult[0]
    prod=getResult[1]
    inv=getResult[2]
    detail=getResult[3]
    return render_template('index.html',cus=cus,prod=prod,inv=inv,detail=detail)

@app.route('/allcustomers',methods=['GET'])
def allCus():
    global data
    with open('./jsonData/customers.json') as json_file:
	    data = json.load(json_file)
    return jsonify(list(data.values())), 200

@app.route('/allproducts',methods=['GET'])
def allProd():
    global data
    with open('./jsonData/products.json') as json_file:
	    data = json.load(json_file)
    return jsonify(list(data.values())), 200

@app.route('/allinvoices',methods=['GET'])
def allInv():
    global data
    with open('./jsonData/invoices.json') as json_file:
	    data = json.load(json_file)
    return jsonify(list(data.values())), 200

@app.route('/alldetail',methods=['GET'])
def allDetail():
    global data
    with open('./jsonData/invdetail.json') as json_file:
	    data = json.load(json_file)
    return jsonify(list(data.values())), 200


app.run(port=5000) 